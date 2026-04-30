from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill

from database import get_db
from auth import get_current_user
import models

router = APIRouter()

HEADER_BG = "1a1a2e"
ALT_ROW_BG = "f8f8ff"


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)


def _fmt_date(d):
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _get_orders(db, current_user, supplier=None, buyer=None, category=None,
                brand=None, from_date=None, to_date=None, po_currency=None):
    q = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.user_id == current_user.id,
        models.PurchaseOrder.status != "deleted",
    )
    if supplier:
        q = q.filter(models.PurchaseOrder.supplier.ilike(f"%{supplier}%"))
    if buyer:
        q = q.filter(models.PurchaseOrder.buyer.ilike(f"%{buyer}%"))
    if category:
        q = q.filter(models.PurchaseOrder.category.ilike(f"%{category}%"))
    if brand:
        q = q.filter(models.PurchaseOrder.brand.ilike(f"%{brand}%"))
    if po_currency:
        q = q.filter(models.PurchaseOrder.po_currency == po_currency.upper())
    if from_date:
        try:
            q = q.filter(models.PurchaseOrder.created_at >= datetime.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        try:
            q = q.filter(models.PurchaseOrder.created_at <= datetime.fromisoformat(to_date))
        except ValueError:
            pass
    return q.order_by(models.PurchaseOrder.created_at.desc()).all()


@router.get("")
def export_data(
    format: str = "excel",
    supplier: Optional[str] = None,
    buyer: Optional[str] = None,
    category: Optional[str] = None,
    brand: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    po_currency: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    orders = _get_orders(db, current_user, supplier, buyer, category,
                         brand, from_date, to_date, po_currency)

    PO_HEADERS = [
        "PO Number", "Business Unit", "Supplier", "Factory", "Brand", "Buyer",
        "Department", "Category", "Style No", "New/Rebuy", "Color", "Country",
        "Supplier Ref No", "Product Description", "PO Received Date",
        "Total Order Qty", "USD Price/PC", "GBP Price/PC",
        "Total Value USD", "Total Value GBP", "PO Currency", "Exchange Rate",
        "Confirmed Ex-Factory", "Revised Ex-Factory", "Delivery Date",
        "Mode", "Port of Loading", "Sample Approved", "Sustainable", "Incoterms",
        "Status",
    ]

    def po_row(po):
        return [
            po.po_number, po.business_unit, po.supplier, po.factory, po.brand, po.buyer,
            po.department, po.category, po.style_number, po.new_rebuy, po.color, po.country,
            po.supplier_ref_no, po.product_description, _fmt_date(po.order_date),
            po.total_order_qty, po.usd_price_per_pc, po.gbp_price_per_pc,
            po.total_value_usd, po.total_value_gbp, po.po_currency, po.exchange_rate,
            _fmt_date(po.confirmed_ex_factory), _fmt_date(po.revised_ex_factory),
            _fmt_date(po.delivery_date),
            po.mode, po.port_of_loading, po.sample_approved_status, po.sustainable,
            po.incoterms, po.status,
        ]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(PO_HEADERS)
        for po in orders:
            writer.writerow(po_row(po))
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=purchase_orders.csv"},
        )

    wb = Workbook()
    ws_po = wb.active
    ws_po.title = "Purchase Orders"

    for col, header in enumerate(PO_HEADERS, 1):
        _style_header(ws_po.cell(row=1, column=col, value=header))

    for row_idx, po in enumerate(orders, 2):
        alt = row_idx % 2 == 0
        for col_idx, val in enumerate(po_row(po), 1):
            cell = ws_po.cell(row=row_idx, column=col_idx, value=val)
            if alt:
                cell.fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
            if col_idx in [17, 18, 19, 20] and isinstance(val, (int, float)):
                cell.number_format = '0.00'

    sr = len(orders) + 2
    ws_po.cell(row=sr, column=1, value="TOTALS").font = Font(bold=True)
    ws_po.cell(row=sr, column=16, value=sum(po.total_order_qty or 0 for po in orders)).font = Font(bold=True)
    ws_po.cell(row=sr, column=19, value=round(sum(po.total_value_usd or 0 for po in orders), 2)).font = Font(bold=True)
    ws_po.cell(row=sr, column=20, value=round(sum(po.total_value_gbp or 0 for po in orders), 2)).font = Font(bold=True)

    _auto_width(ws_po)

    ws_li = wb.create_sheet("Line Items")
    LI_HEADERS = [
        "PO Number", "Supplier", "Brand", "Style Number",
        "Order Qty", "USD Price/PC", "GBP Price/PC",
        "Line Total USD", "Line Total GBP",
        "Confirmed Delivery", "Actual Delivery",
    ]
    for col, header in enumerate(LI_HEADERS, 1):
        _style_header(ws_li.cell(row=1, column=col, value=header))

    li_row = 2
    for po in orders:
        for item in po.line_items:
            usd_pc = po.usd_price_per_pc or item.unit_price or 0
            gbp_pc = po.gbp_price_per_pc or 0
            qty = item.order_quantity or 0
            cell1 = ws_li.cell(row=li_row, column=1, value=po.po_number)
            ws_li.cell(row=li_row, column=2, value=po.supplier)
            ws_li.cell(row=li_row, column=3, value=po.brand)
            ws_li.cell(row=li_row, column=4, value=item.style_number)
            ws_li.cell(row=li_row, column=5, value=qty)
            ws_li.cell(row=li_row, column=6, value=usd_pc).number_format = '0.00'
            ws_li.cell(row=li_row, column=7, value=gbp_pc).number_format = '0.00'
            ws_li.cell(row=li_row, column=8, value=item.line_total_usd or round(qty * usd_pc, 2)).number_format = '0.00'
            ws_li.cell(row=li_row, column=9, value=item.line_total_gbp or round(qty * gbp_pc, 2)).number_format = '0.00'
            ws_li.cell(row=li_row, column=10, value=_fmt_date(item.delivery_date_confirmed))
            ws_li.cell(row=li_row, column=11, value=_fmt_date(item.delivery_date_actual))
            li_row += 1

    _auto_width(ws_li)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_orders.xlsx"},
    )
